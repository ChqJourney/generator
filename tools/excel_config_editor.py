#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel批量配置编辑器

将提取的元素导出为Excel，方便批量编辑，然后再转回JSON

用法:
    # 导出到Excel
    python tools/excel_config_editor.py export config/extracted_elements.json --output config/fields.xlsx
    
    # 从Excel导入
    python tools/excel_config_editor.py import config/fields.xlsx --output config/report_config.json
"""

import json
import argparse
import sys
import os
from typing import List, Dict

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装openpyxl: pip install openpyxl")
    sys.exit(1)


def export_to_excel(input_json: str, output_excel: str):
    """
    将提取的元素导出为Excel表格
    """
    # 读取JSON
    with open(input_json, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Mappings"
    
    # 定义表头
    headers = [
        "序号",
        "Section No (节号)",
        "Template Field (占位符)",
        "Type (类型)",
        "Source Prefix (数据源)",
        "Source Field (字段名)",
        "Function (计算函数)",
        "Args (参数)",
        "Location (位置)",
        "Extra Config (额外配置)"
    ]
    
    # 写入表头
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 准备数据
    mappings = data.get("field_mappings", [])
    if not mappings:
        # 如果没有现成的mappings，从placeholders和checkboxes生成
        from extract_template_elements import generate_field_mappings
        mappings = generate_field_mappings(
            data.get("placeholders", {}),
            data.get("checkboxes", [])
        )
    
    # 按section_no排序
    mappings.sort(key=lambda x: x.get("section_no", 1))
    
    # 写入数据
    type_colors = {
        "text": "E2EFDA",      # 浅绿
        "checkbox": "FFF2CC",   # 浅黄
        "table": "DDEBF7",      # 浅蓝
        "image": "FCE4D6"       # 浅橙
    }
    
    for row, mapping in enumerate(mappings, 2):
        field_type = mapping.get("type", "text")
        fill_color = type_colors.get(field_type, "FFFFFF")
        
        # 分解source_field，只取字段名（去掉前缀）
        source_field_full = mapping.get("source_field", "")
        if "." in source_field_full:
            source_prefix, field_name = source_field_full.split(".", 1)
        else:
            source_prefix = source_field_full
            field_name = source_field_full
        
        # 准备额外配置
        extra_config = []
        if field_type == "table":
            extra = []
            if mapping.get("table_template_path"):
                extra.append(f"template:{mapping['table_template_path']}")
            if mapping.get("row_strategy"):
                extra.append(f"strategy:{mapping['row_strategy']}")
            if mapping.get("header_rows") is not None:
                extra.append(f"headers:{mapping['header_rows']}")
            extra_config = extra
        elif field_type == "image":
            extra = []
            if mapping.get("width"):
                extra.append(f"width:{mapping['width']}")
            if mapping.get("alignment"):
                extra.append(f"align:{mapping['alignment']}")
            extra_config = extra
        
        row_data = [
            row - 1,  # 序号
            mapping.get("section_no", 1),  # Section No
            mapping.get("template_field", ""),
            field_type,
            source_prefix,
            field_name,  # 只显示字段名，不含前缀
            mapping.get("function", ""),
            ", ".join(mapping.get("args", [])),
            mapping.get("location", "body"),
            "; ".join(extra_config)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # 设置列宽
    column_widths = [6, 12, 25, 10, 15, 25, 20, 30, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 添加说明sheet
    help_ws = wb.create_sheet("使用说明")
    help_content = [
        ["字段配置说明"],
        [],
        ["Section No (节号)", "Word文档中的节号(section)，从1开始"],
        ["Type (类型)", "可选值: text, checkbox, table, image"],
        ["Source Prefix (数据源)", "可选值: metadata, extracted_data, calculated_data"],
        ["Source Field (字段名)", "字段名称（不含前缀）"],
        ["Function (计算函数)", "calculator.py中注册的函数名（计算字段需填写）"],
        ["Args (参数)", "逗号分隔的参数路径，如: extracted_data.wattage, extracted_data.flux"],
        ["Extra Config (额外配置)", "table: template:xxx.docx;strategy:fixed_rows;headers:1"],
        ["", "image: width:4.0;align:center"],
        [],
        ["常用计算函数"],
        ["calculate_energy_class_rating", "计算能效等级，args: wattage, flux"],
        ["calculate_energy_efficacy", "计算能效值，args: wattage, flux"],
        ["calculate_percentage", "计算百分比，args: value, total"],
        ["format_number", "格式化数字，args: value, decimal"],
    ]
    
    for row, line in enumerate(help_content, 1):
        if isinstance(line, list):
            for col, value in enumerate(line, 1):
                cell = help_ws.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = Font(bold=True, size=14)
                elif col == 1:
                    cell.font = Font(bold=True)
    
    help_ws.column_dimensions["A"].width = 35
    help_ws.column_dimensions["B"].width = 60
    
    # 保存
    wb.save(output_excel)
    print(f"已导出到Excel: {output_excel}")
    print(f"共 {len(mappings)} 个字段")


def import_from_excel(input_excel: str, output_json: str):
    """
    从Excel导入配置并生成report_config.json
    """
    # 读取Excel
    wb = load_workbook(input_excel)
    ws = wb["Field Mappings"]
    
    mappings = []
    
    # 从第2行开始读取（跳过表头）
    for row in range(2, ws.max_row + 1):
        template_field = ws.cell(row=row, column=3).value  # Template Field在第3列
        if not template_field:
            continue
        
        section_no = ws.cell(row=row, column=2).value or 1  # Section No在第2列
        field_type = ws.cell(row=row, column=4).value or "text"  # Type在第4列
        source_prefix = ws.cell(row=row, column=5).value or "extracted_data"  # Source Prefix在第5列
        field_name = ws.cell(row=row, column=6).value or template_field  # Source Field在第6列
        source_field = f"{source_prefix}.{field_name}"  # 重新组装完整路径
        function = ws.cell(row=row, column=7).value or ""  # Function在第7列
        args_str = ws.cell(row=row, column=8).value or ""  # Args在第8列
        location = ws.cell(row=row, column=9).value or "body"  # Location在第9列
        extra_config = ws.cell(row=row, column=10).value or ""  # Extra Config在第10列
        
        mapping = {
            "template_field": template_field,
            "source_field": source_field,
            "type": field_type,
            "section_no": section_no
        }
        
        # 添加计算字段属性（如果填写了function）
        if function:
            mapping["function"] = function
            if args_str:
                mapping["args"] = [a.strip() for a in str(args_str).split(",") if a.strip()]
        
        # 根据类型添加额外配置
        if field_type == "table":
            # 解析extra_config
            if extra_config:
                configs = {}
                for part in str(extra_config).split(";"):
                    if ":" in part:
                        key, value = part.strip().split(":", 1)
                        configs[key.strip()] = value.strip()
                
                if "template" in configs:
                    mapping["table_template_path"] = configs["template"]
                if "strategy" in configs:
                    mapping["row_strategy"] = configs["strategy"]
                if "headers" in configs:
                    mapping["header_rows"] = int(configs["headers"])
        
        elif field_type == "image":
            if extra_config:
                configs = {}
                for part in str(extra_config).split(";"):
                    if ":" in part:
                        key, value = part.strip().split(":", 1)
                        configs[key.strip()] = value.strip()
                
                if "width" in configs:
                    mapping["width"] = float(configs["width"])
                if "align" in configs:
                    mapping["alignment"] = configs["align"]
        
        mappings.append(mapping)
    
    # 构建最终配置
    config = {
        "template_path": "",  # 需要手动填写
        "output_dir": "output/",
        "field_mappings": mappings,
        "merge_strategy": "prefer_first"
    }
    
    # 保存
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    
    print(f"已生成配置: {output_json}")
    print(f"共 {len(mappings)} 个字段")
    
    # 列出计算字段
    calculated = [m for m in mappings if m.get("function")]
    if calculated:
        print(f"\n包含 {len(calculated)} 个计算字段:")
        for m in calculated:
            print(f"  - {m['template_field']}: {m['function']}")


def main():
    parser = argparse.ArgumentParser(
        description="Excel配置编辑器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    # 导出到Excel
    python tools/excel_config_editor.py export config/extracted.json --output fields.xlsx
    
    # 从Excel导入
    python tools/excel_config_editor.py import fields.xlsx --output report_config.json
        """
    )
    
    subparsers = parser.add_subparsers(dest="command", help="可用命令")
    
    # export命令
    export_parser = subparsers.add_parser("export", help="导出到Excel")
    export_parser.add_argument("input", help="输入JSON文件")
    export_parser.add_argument("--output", "-o", default="config/fields.xlsx", 
                              help="输出Excel文件")
    
    # import命令
    import_parser = subparsers.add_parser("import", help="从Excel导入")
    import_parser.add_argument("input", help="输入Excel文件")
    import_parser.add_argument("--output", "-o", default="config/report_config.json",
                              help="输出JSON文件")
    
    args = parser.parse_args()
    
    if args.command == "export":
        export_to_excel(args.input, args.output)
    elif args.command == "import":
        import_from_excel(args.input, args.output)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
