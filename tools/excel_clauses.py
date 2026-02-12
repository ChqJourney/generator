"""
Excel 条款配置转换工具
支持 Excel <-> JSON 双向转换
"""

import json
import sys
import argparse
import os
from pathlib import Path
from typing import Dict, List, Any, Optional

sys.path.insert(0, str(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from src.utils.logging_config import get_logger

logger = get_logger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.error("需要安装 openpyxl。请运行：pip install openpyxl")
    sys.exit(1)

# 尝试导入 json5 以支持 JSONC 格式
try:
    import json5
    HAS_JSON5 = True
except ImportError:
    HAS_JSON5 = False


def parse_args_mapping(args_str: str) -> List[str]:
    """解析 args 字段，支持逗号分隔或JSON格式"""
    args_str = args_str.strip()
    if args_str.startswith('[') and args_str.endswith(']'):
        try:
            return json.loads(args_str)
        except:
            pass
    return [a.strip() for a in args_str.split(',') if a.strip()]


def parse_rules(rules_str: str) -> List[Dict[str, str]]:
    """解析 rules 字段"""
    rules_str = rules_str.strip()
    if not rules_str:
        return []
    
    if rules_str.startswith('['):
        try:
            return json.loads(rules_str)
        except:
            pass
    
    rules = []
    for rule_part in rules_str.split(';'):
        rule_part = rule_part.strip()
        if '|' in rule_part:
            condition, result = rule_part.split('|', 1)
            rules.append({
                "condition": condition.strip(),
                "result": result.strip()
            })
    return rules


def excel_to_json(excel_path: str, json_path: str, config_path: Optional[str] = None):
    """将 Excel 转换为 JSON 配置"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    field_mappings = []
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        try:
            row_data = dict(zip(headers, row))
            
            clause_id = row_data.get('clause_id', '').strip()
            if not clause_id:
                continue
            
            template_field = row_data.get('template_field', '').strip() or clause_id
            param_names_str = row_data.get('param_names', '')
            args_str = row_data.get('args', '')
            rules_str = row_data.get('rules', '')
            default = row_data.get('default', 'N/A')
            description = row_data.get('description', '')
            
            param_names = [p.strip() for p in param_names_str.split(',') if p.strip()]
            args = parse_args_mapping(args_str)
            rules = parse_rules(rules_str)
            
            if not rules:
                errors.append(f"第{row_idx}行 ({clause_id}): 没有有效的规则")
                continue
            
            field_mapping = {
                "template_field": template_field,
                "source_field": f"calculated_data.{clause_id}",
                "type": "text",
                "function": "evaluate_clause",
                "args": args,
                "clause_config": {
                    "clause_id": clause_id,
                    "param_names": param_names,
                    "rules": rules,
                    "default": default
                }
            }
            
            if description:
                field_mapping["clause_config"]["description"] = description
            
            field_mappings.append(field_mapping)
            
        except Exception as e:
            errors.append(f"第{row_idx}行: {e}")
    
    output = {"field_mappings": field_mappings}
    
    if config_path and Path(config_path).exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            original_config = json.load(f)
            for key in original_config:
                if key != 'field_mappings':
                    output[key] = original_config[key]
            for mapping in original_config.get('field_mappings', []):
                if mapping.get('function') != 'evaluate_clause':
                    output['field_mappings'].append(mapping)
    
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    logger.info(f"成功生成 {len(field_mappings)} 个条款配置")
    logger.info(f"输出文件: {json_path}")

    if errors:
        logger.warning(f"警告 ({len(errors)} 个问题):")
        for error in errors[:10]:
            logger.warning(f"  - {error}")
        if len(errors) > 10:
            logger.warning(f"  ... 还有 {len(errors) - 10} 个问题")


def load_json_with_comments(file_path: str) -> dict:
    """加载可能包含注释的 JSON 文件"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 尝试使用 json5 解析
    if HAS_JSON5:
        return json5.loads(content)
    
    # 回退到标准 json（仅支持无注释的 JSON）
    return json.loads(content)


def json_to_excel(json_path: str, excel_path: str):
    """将 JSON 配置转换为 Excel（反向导出）"""
    config = load_json_with_comments(json_path)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "条款配置"
    
    headers = [
        'clause_id',
        'template_field', 
        'param_names',
        'args',
        'rules',
        'default',
        'description'
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 2
    clause_count = 0
    
    for mapping in config.get('field_mappings', []):
        if mapping.get('function') != 'evaluate_clause':
            continue
        
        clause_config = mapping.get('clause_config', {})
        clause_id = clause_config.get('clause_id', '')
        
        if not clause_id:
            continue
        
        rules = clause_config.get('rules', [])
        rules_str = '; '.join([
            f"{r.get('condition', '')}|{r.get('result', '')}"
            for r in rules
        ])
        
        data = {
            'clause_id': clause_id,
            'template_field': mapping.get('template_field', clause_id),
            'param_names': ', '.join(clause_config.get('param_names', [])),
            'args': ', '.join(mapping.get('args', [])),
            'rules': rules_str,
            'default': clause_config.get('default', 'N/A'),
            'description': clause_config.get('description', '')
        }
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=data.get(header, ''))
        
        row += 1
        clause_count += 1
    
    column_widths = {
        'A': 15, 'B': 15, 'C': 35, 'D': 50, 'E': 80, 'F': 10, 'G': 40
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.row_dimensions[1].height = 25
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_cells in ws[f'A1:G{row-1}']:
        for cell in row_cells:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    wb.save(excel_path)

    logger.info(f"成功导出 {clause_count} 个条款配置")
    logger.info(f"Excel文件: {excel_path}")


def create_template(excel_path: str):
    """创建 Excel 模板文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "条款配置"
    
    headers = [
        'clause_id',
        'template_field', 
        'param_names',
        'args',
        'rules',
        'default',
        'description'
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    examples = [
        [
            'v.1.a',
            'v.1.a',
            'containing_product,light_sources',
            'metadata.containing_product.value,metadata.light_sources.value',
            "containing_product == 'true' AND light_sources == 'true'|Pass;containing_product == 'false'|N/A",
            'Fail',
            'Containing product should have at least one light source'
        ],
        [
            'v.1.b.1',
            'v.1.b.1',
            'replaceable_LED,replace_by_pro,non_replaceable',
            'metadata.replaceable_LED.value,metadata.replaceable_by_professional.value,metadata.non_replaceable.value',
            "replaceable_LED == 'true'|Pass;non_replaceable == 'true'|N/A;replaceable_LED == 'false'|N/A",
            'Fail',
            'Replaceable LED light source check'
        ],
        [
            'v.II.1.a',
            'v.II.1.a',
            'flux,power,non_directional,LED_source',
            'extracted_data.useful_luminous_flux,extracted_data.Pon,metadata.non_directional.value,metadata.LED_source.value',
            "LED_source == 'false'|N/A;flux IS NULL OR power IS NULL|Fail;non_directional == 'true' AND (flux / power) >= 85|Pass;non_directional == 'false' AND (flux / power) >= 75|Pass",
            'Fail',
            'Energy efficiency class calculation'
        ]
    ]
    
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    column_widths = {
        'A': 15, 'B': 15, 'C': 35, 'D': 60, 'E': 100, 'F': 10, 'G': 45
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.row_dimensions[1].height = 25
    
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for row in range(2, 2 + len(examples)):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = example_fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    # 添加说明工作表
    ws_help = wb.create_sheet("使用说明")
    help_content = [
        ['条款配置 Excel 使用说明', ''],
        ['', ''],
        ['列说明：', ''],
        ['clause_id', '条款唯一标识，如 v.1.a、v.II.1.a'],
        ['template_field', 'Word模板中的占位符名称（可选）'],
        ['param_names', '参数名列表，用逗号分隔。这些名称在 rules 中使用'],
        ['args', '数据源路径，对应 param_names，如 metadata.containing_product.value'],
        ['rules', '规则列表，使用简单格式：条件|结果;条件|结果'],
        ['default', '默认返回值，当没有规则匹配时使用'],
        ['description', '条款描述（可选）'],
        ['', ''],
        ['rules 格式说明：', ''],
        ['简单格式：', 'condition1|result1;condition2|result2'],
        ['示例：', "containing_product == 'true' AND light_sources == 'true'|Pass;containing_product == 'false'|N/A"],
        ['', ''],
        ['支持的条件语法：', ''],
        ['比较：', '==、!=、>、<、>=、<='],
        ['逻辑：', 'AND、OR、NOT'],
        ['特殊：', 'IN、NOT IN、CONTAINS、IS NULL、IS NOT NULL'],
        ['分组：', '使用括号 () 分组'],
        ['', ''],
        ['使用步骤：', ''],
        ['1.', '在模板基础上填写或修改条款配置'],
        ['2.', '保存 Excel 文件'],
        ['3.', '运行转换命令：python tools/excel_clauses.py excel2json clauses.xlsx output.json'],
        ['4.', '生成的 JSON 文件可以合并到 report_config.jsonc 中使用'],
        ['', ''],
        ['反向导出：', ''],
        ['', '如需将现有配置导出到 Excel 编辑：'],
        ['', 'python tools/excel_clauses.py json2excel report_config.jsonc clauses.xlsx']
    ]
    
    for row_idx, row_data in enumerate(help_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif value and str(value).endswith('：'):
                cell.font = Font(bold=True)
    
    ws_help.column_dimensions['A'].width = 20
    ws_help.column_dimensions['B'].width = 80
    
    wb.save(excel_path)
    logger.info(f"已创建模板文件: {excel_path}")
    logger.info(f"包含 {len(examples)} 个示例条款")


def main():
    parser = argparse.ArgumentParser(
        description='Excel 条款配置转换工具 - 支持 Excel 与 JSON 双向转换'
    )
    
    subparsers = parser.add_subparsers(dest='command', help='可用命令')
    
    # excel2json 命令
    parser_excel2json = subparsers.add_parser('excel2json', help='将 Excel 转换为 JSON')
    parser_excel2json.add_argument('excel', help='输入 Excel 文件路径')
    parser_excel2json.add_argument('output', help='输出 JSON 文件路径')
    parser_excel2json.add_argument('--config', '-c', help='原始配置文件路径（用于合并非条款配置）')
    
    # json2excel 命令
    parser_json2excel = subparsers.add_parser('json2excel', help='将 JSON 转换为 Excel')
    parser_json2excel.add_argument('json', help='输入 JSON 文件路径')
    parser_json2excel.add_argument('output', help='输出 Excel 文件路径')
    
    # template 命令
    parser_template = subparsers.add_parser('template', help='创建 Excel 模板文件')
    parser_template.add_argument('output', help='输出 Excel 文件路径')
    
    # validate 命令
    parser_validate = subparsers.add_parser('validate', help='验证 Excel 配置')
    parser_validate.add_argument('excel', help='输入 Excel 文件路径')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    if args.command == 'excel2json':
        excel_to_json(args.excel, args.output, args.config)
    elif args.command == 'json2excel':
        json_to_excel(args.json, args.output)
    elif args.command == 'template':
        create_template(args.output)
    elif args.command == 'validate':
        validate_excel(args.excel)


def validate_excel(excel_path: str):
    """验证 Excel 配置"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    errors = []
    warnings = []
    valid_count = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        row_data = dict(zip(headers, row))
        clause_id = row_data.get('clause_id', '').strip()
        
        if not clause_id:
            continue
        
        # 验证必需字段
        if not row_data.get('param_names', '').strip():
            errors.append(f"第{row_idx}行 ({clause_id}): 缺少 param_names")
        
        if not row_data.get('rules', '').strip():
            errors.append(f"第{row_idx}行 ({clause_id}): 缺少 rules")
        
        # 验证 rules 格式
        rules_str = row_data.get('rules', '')
        if rules_str.strip():
            rules = parse_rules(rules_str)
            if not rules:
                errors.append(f"第{row_idx}行 ({clause_id}): rules 格式无效")
            else:
                # 验证每个规则的 condition
                for i, rule in enumerate(rules):
                    condition = rule.get('condition', '')
                    if not condition:
                        warnings.append(f"第{row_idx}行 ({clause_id}) 规则{i+1}: condition 为空")
        
        valid_count += 1
    
    logger.info(f"验证完成")
    logger.info(f"有效条款数: {valid_count}")

    if errors:
        logger.error(f"错误 ({len(errors)}):")
        for error in errors[:10]:
            logger.error(f"  - {error}")
        if len(errors) > 10:
            logger.error(f"  ... 还有 {len(errors) - 10} 个错误")

    if warnings:
        logger.warning(f"警告 ({len(warnings)}):")
        for warning in warnings[:5]:
            logger.warning(f"  - {warning}")
        if len(warnings) > 5:
            logger.warning(f"  ... 还有 {len(warnings) - 5} 个警告")

    if not errors:
        logger.info("验证通过！配置格式正确。")
    else:
        logger.error(f"发现 {len(errors)} 个错误，请修正后再转换。")


if __name__ == '__main__':
    main()
