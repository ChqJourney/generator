"""
将字段映射配置转换为processor.py操作队列
新架构：从calculated_report.json读取数据，支持点号路径访问
"""
import json
import sys
import re
from pathlib import Path
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from utils.path_navigator import DataNavigator


def load_json(path: str) -> Dict:
    """加载JSON文件"""
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_value_by_path(data: Dict, path: str) -> Any:
    """
    通过点号路径获取值
    
    Args:
        data: 分层数据字典
        path: 点号分隔的路径，如 'extracted_data.rated_wattage'
        
    Returns:
        路径对应的值
    """
    return DataNavigator.get_value(data, path)


def is_external_table_reference(value: Any) -> bool:
    """
    判断值是否为外部表格引用（需要读取Excel）
    
    Args:
        value: 字段值
        
    Returns:
        bool: 是否需要从外部Excel读取
    """
    return (
        isinstance(value, dict) and 
        value.get('type') == 'external' and 
        'source_id' in value
    )


def is_direct_table_data(value: Any) -> bool:
    """
    判断值是否为直接的表格数据（列表的列表）
    
    Args:
        value: 字段值
        
    Returns:
        bool: 是否是直接表格数据
    """
    return (
        isinstance(value, list) and 
        len(value) > 0 and 
        isinstance(value[0], list)
    )


def generate_operations(config: Dict, report_data: Dict) -> Dict:
    """
    生成操作队列
    
    Args:
        config: 配置字典
        report_data: 计算后的报告数据（calculated_report.json）
        
    Returns:
        Dict: 操作队列
    """
    operations = []
    
    for mapping in config.get('field_mappings', []):
        source_field = mapping.get('source_field')
        field_type = mapping.get('type')
        placeholder = mapping.get('template_field')
        
        if not source_field:
            print(f"Warning: Missing source_field for {placeholder}, skipping")
            continue
        
        # 通过点号路径获取值
        value = get_value_by_path(report_data, source_field)
        
        if value is None:
            print(f"Warning: Value not found for path '{source_field}', skipping {placeholder}")
            continue
        
        if field_type == 'text':
            operations.append({
                'type': 'text',
                'placeholder': placeholder,
                'value': str(value)
            })
        
        elif field_type == 'image':
            image_paths = value if isinstance(value, list) else [value]
            if image_paths and isinstance(image_paths[0], str):
                try:
                    parsed = json.loads(image_paths[0])
                    image_paths = parsed
                except:
                    pass
            operations.append({
                'type': 'image',
                'placeholder': placeholder,
                'image_paths': image_paths,
                'width': mapping.get('width'),
                'height': mapping.get('height'),
                'alignment': mapping.get('alignment')
            })
        
        elif field_type == 'table':
            table_data = None
            
            # 判断数据源类型
            if is_direct_table_data(value):
                # 直接是表格数据（内嵌）
                table_data = value
                print(f"Using embedded table data for {placeholder}")
            elif is_external_table_reference(value):
                # 外部Excel引用
                target_headers = mapping.get('target_headers')
                table_data = build_table_data_from_excel(value, target_headers)
                print(f"Loaded table data from Excel for {placeholder}")
            else:
                print(f"Warning: Unrecognized table data format for {placeholder}")
                continue
            
            operation = {
                'type': 'table',
                'placeholder': placeholder,
                'table_template_path': mapping['table_template_path'],
                'table_data': table_data
            }
            
            # 添加可选参数
            for key in ['transformations', 'row_strategy', 'skip_columns', 'header_rows']:
                if key in mapping:
                    operation[key] = mapping[key]
            
            operations.append(operation)
    
    return {'operations': operations}


def build_table_data_from_excel(value: Dict, target_headers: Optional[List[str]] = None) -> List[List[str]]:
    """
    从Excel文件构建表格数据
    
    Args:
        value: 包含source_id等信息的字典
        target_headers: 目标表头列表
        
    Returns:
        List[List[str]]: 表格数据
    """
    source_id = value['source_id']
    sources = source_id.split('|')
    if len(sources) != 2:
        print(f"Error: Invalid source_id format '{source_id}', expected 'file.xlsx|SheetName'")
        return []
    
    file_path = sources[0]
    sheet_name = sources[1]
    start_row = value.get('start_row', 0)
    mapping = value.get('mapping', {})
    actual_path = 'data_files/' + file_path
    
    print(f"Building table data from file: {actual_path}, sheet: {sheet_name}, start_row: {start_row}")
    
    try:
        table_data = get_xlsx_to_list(actual_path, sheet_name, start_row, mapping, target_headers)
        print(f"Extracted {len(table_data)} rows of table data")
        return table_data
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []


def get_xlsx_to_list(file_path: str, sheet_name: str, start_row: int, 
                     mapping: Dict, target_headers: Optional[List[str]]) -> List[List[str]]:
    """
    读取Excel文件并转换为列表格式
    """
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    
    try:
        ws = wb[sheet_name]
    except KeyError:
        wb.close()
        raise ValueError(f"工作表 '{sheet_name}' 不存在")

    header_row_num = start_row + 1
    row_iterator = ws.iter_rows(min_row=header_row_num, values_only=True)
    
    try:
        header_values = next(row_iterator)
    except StopIteration:
        wb.close()
        return []
    
    def normalize_header(h):
        if h is None:
            return ""
        return " ".join(str(h).split())
    
    excel_headers_map = {}
    if header_values:
        for idx, val in enumerate(header_values):
            if val is not None:
                excel_headers_map[normalize_header(val)] = idx
    
    target_col_indices = []
    
    if isinstance(mapping, list):
        mapping_dict = {m.get('name'): m.get('mapColumn') for m in mapping if isinstance(m, dict)}
    else:
        mapping_dict = mapping
    
    if target_headers:
        for header in target_headers:
            source_col_name = mapping_dict.get(header)
            col_idx = None
            if source_col_name:
                normalized_source = normalize_header(source_col_name)
                if normalized_source in excel_headers_map:
                    col_idx = excel_headers_map[normalized_source]
                else:
                    print(f"警告: 未找到映射列 '{source_col_name}' (规范化: '{normalized_source}')", file=sys.stderr)
            target_col_indices.append(col_idx)
    
    data = []
    for row in row_iterator:
        row_data = []
        for col_idx in target_col_indices:
            if col_idx is not None and col_idx < len(row):
                cell_val = row[col_idx]
                row_data.append(str(cell_val) if cell_val is not None else "")
            else:
                row_data.append("")
        data.append(row_data)
    
    wb.close()
    return data


def main():
    """命令行入口"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Generate processor operations from field mappings (new architecture)'
    )
    parser.add_argument(
        '--config', 
        required=True, 
        help='Path to report_config.json'
    )
    parser.add_argument(
        '--report', 
        required=True, 
        help='Path to calculated_report.json'
    )
    parser.add_argument(
        '--output', 
        required=True, 
        help='Path to output operations.json'
    )
    
    args = parser.parse_args()
    
    try:
        # 加载数据
        config = load_json(args.config)
        report_data = load_json(args.report)
        
        # 生成操作队列
        operations = generate_operations(config, report_data)
        
        # 保存结果
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(operations, f, indent=2, ensure_ascii=False)
        
        op_count = len(operations.get('operations', []))
        print(f"Generated {op_count} operations: {args.output}")
        return 0
    
    except FileNotFoundError as e:
        print(f"Error: File not found - {e}", file=sys.stderr)
        return 1
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON - {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
